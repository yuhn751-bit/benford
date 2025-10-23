#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
벤포드 분석 웹 애플리케이션 (Streamlit)
첫 자리(1-9) 분석 버전
"""

import streamlit as st
import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import io
from datetime import datetime

# 페이지 설정
st.set_page_config(
    page_title="벤포드 분석 도구",
    page_icon="🔍",
    layout="wide"
)


class BenfordAnalyzer:
    """벤포드 분석 클래스"""
    
    def __init__(self, file_data, confidence=0.95):
        self.file_data = file_data
        self.confidence = confidence
        self.z_critical = stats.norm.ppf((1 + confidence) / 2)
        self.sigma_thresholds = [1.0, 1.5, 2.0]
        
        self.wb = None
        self.data = None
        self.keywords = None
        self.results = {}
        
    def benford_probability(self, digit):
        """벤포드 법칙 확률 계산"""
        return np.log10(1 + 1 / digit)
    
    def leading_digit(self, number):
        """숫자의 첫 자리 추출"""
        number = abs(number)
        if number == 0:
            return None
        
        while number >= 10:
            number /= 10
        
        return int(number)
    
    def load_data(self):
        """엑셀 데이터 로딩"""
        self.wb = load_workbook(self.file_data)
        
        # 분개장 데이터
        df = pd.read_excel(self.file_data, sheet_name='분개장입력')
        
        # 금액 숫자 변환
        df['금액'] = pd.to_numeric(df['금액'], errors='coerce')
        df = df.dropna(subset=['금액'])
        
        # 첫 자리 추출
        df['첫자리'] = df['금액'].apply(self.leading_digit)
        df = df[(df['첫자리'] >= 1) & (df['첫자리'] <= 9)]
        
        self.data = df
        
        # 키워드 로딩
        try:
            keywords_df = pd.read_excel(self.file_data, sheet_name='적요 키워드')
            self.keywords = keywords_df.iloc[:, 0].dropna().astype(str).str.strip().tolist()
        except:
            self.keywords = []
    
    def analyze_benford(self, data, label="전체"):
        """벤포드 분석 수행"""
        n = len(data)
        if n == 0:
            return None
        
        results = []
        
        for digit in range(1, 10):
            observed = (data['첫자리'] == digit).sum()
            expected_prob = self.benford_probability(digit)
            expected_count = expected_prob * n
            
            obs_ratio = observed / n if n > 0 else 0
            std_dev = np.sqrt(n * expected_prob * (1 - expected_prob))
            z_value = (observed - expected_count) / std_dev if std_dev > 0 else 0
            chi_sq = (observed - expected_count) ** 2 / expected_count if expected_count > 0 else 0
            
            flag = "초과" if abs(z_value) >= self.z_critical else ""
            
            results.append({
                '구분': label,
                '첫자리': digit,
                '관측빈도': observed,
                '관측비율': obs_ratio,
                '기대비율': expected_prob,
                '기대빈도': expected_count,
                '(O-E)^2/E': chi_sq,
                '표본수': n,
                '표준편차': std_dev,
                'Z값': z_value,
                '신뢰수준판정': flag
            })
        
        return pd.DataFrame(results)
    
    def run_analysis(self):
        """전체 분석 실행"""
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # 1. 전체 분석
        status_text.text("1/8 전체 데이터 분석 중...")
        progress_bar.progress(12)
        self.results['01_벤포드_전체'] = self.analyze_benford(self.data, "전체")
        
        # 2. 손익별
        if '손익' in self.data.columns:
            status_text.text("2/8 손익별 분석 중...")
            progress_bar.progress(25)
            dfs = []
            for pnl in self.data['손익'].unique():
                if pd.notna(pnl):
                    df = self.analyze_benford(self.data[self.data['손익'] == pnl], pnl)
                    if df is not None:
                        dfs.append(df)
            if dfs:
                self.results['02_벤포드_손익'] = pd.concat(dfs, ignore_index=True)
        
        # 3. 대분류별
        if '대분류' in self.data.columns:
            status_text.text("3/8 대분류별 분석 중...")
            progress_bar.progress(37)
            dfs = []
            for cat in self.data['대분류'].unique():
                if pd.notna(cat):
                    df = self.analyze_benford(self.data[self.data['대분류'] == cat], cat)
                    if df is not None:
                        dfs.append(df)
            if dfs:
                self.results['03_벤포드_대분류'] = pd.concat(dfs, ignore_index=True)
        
        # 4. 중분류별
        if '중분류' in self.data.columns:
            status_text.text("4/8 중분류별 분석 중...")
            progress_bar.progress(50)
            dfs = []
            for cat in self.data['중분류'].unique():
                if pd.notna(cat):
                    df = self.analyze_benford(self.data[self.data['중분류'] == cat], cat)
                    if df is not None:
                        dfs.append(df)
            if dfs:
                self.results['04_벤포드_중분류'] = pd.concat(dfs, ignore_index=True)
        
        # 5. 소분류별
        if '소분류' in self.data.columns:
            status_text.text("5/8 소분류별 분석 중...")
            progress_bar.progress(62)
            dfs = []
            for cat in self.data['소분류'].unique():
                if pd.notna(cat):
                    df = self.analyze_benford(self.data[self.data['소분류'] == cat], cat)
                    if df is not None:
                        dfs.append(df)
            if dfs:
                self.results['05_벤포드_소분류'] = pd.concat(dfs, ignore_index=True)
        
        # 6. 일자별 요약 (손익 구분)
        status_text.text("6/8 일자별 요약 중...")
        progress_bar.progress(75)
        if '일자' in self.data.columns and '손익' in self.data.columns:
            daily = self.data.groupby(['일자', '손익']).agg({
                '금액': ['count', 'sum']
            }).reset_index()
            daily.columns = ['일자', '손익', '전표갯수', '금액합계']
            self.results['06_일자별요약'] = daily
        
        # 7. 거래처별 요약 (손익 구분)
        status_text.text("7/8 거래처별 요약 중...")
        progress_bar.progress(87)
        if '거래처' in self.data.columns and '손익' in self.data.columns:
            vendor = self.data.groupby(['거래처', '손익']).agg({
                '금액': ['count', 'sum']
            }).reset_index()
            vendor.columns = ['거래처', '손익', '전표갯수', '금액합계']
            vendor = vendor.sort_values('금액합계', ascending=False)
            self.results['07_거래처별요약'] = vendor
        
        # 8. 키워드 필터링
        if self.keywords:
            status_text.text("8/8 키워드 필터링 중...")
            progress_bar.progress(100)
            keyword_data = []
            for keyword in self.keywords:
                mask = self.data['적요'].astype(str).str.contains(keyword, case=False, na=False)
                matched = self.data[mask].copy()
                if len(matched) > 0:
                    matched.insert(0, '키워드', keyword)
                    keyword_data.append(matched)
            
            if keyword_data:
                cols = ['키워드', '일자', '손익', '금액', '거래처', '적요']
                if '전표번호' in self.data.columns:
                    cols.insert(2, '전표번호')
                if '대분류' in self.data.columns:
                    cols.insert(-3, '대분류')
                if '중분류' in self.data.columns:
                    cols.insert(-3, '중분류')
                if '소분류' in self.data.columns:
                    cols.insert(-3, '소분류')
                
                available_cols = [col for col in cols if col in pd.concat(keyword_data).columns]
                self.results['08_키워드필터'] = pd.concat(keyword_data, ignore_index=True)[available_cols]
        
        # 9. 중점검토 시나리오 (Top 30)
        self.results['09_중점검토'] = self.create_scenario_sheet()
        
        # 10. 분석 해설 (세무조사 내용 강화)
        self.results['10_분석해설'] = self.create_explanation_sheet()
        
        progress_bar.progress(100)
        status_text.text("✅ 분석 완료!")
    
    def create_scenario_sheet(self):
        """중점검토 시나리오 생성 (Top 30)"""
        top_findings = []
        
        for key, df in self.results.items():
            if key.startswith('0') and '벤포드' in key:
                if 'Z값' in df.columns:
                    df_sorted = df.nlargest(30, 'Z값', keep='all')
                    df_sorted['범위'] = key.replace('_벤포드_', ' - ').replace('0', '').replace('_', '')
                    top_findings.append(df_sorted)
        
        if top_findings:
            all_findings = pd.concat(top_findings, ignore_index=True)
            all_findings['Z값_절대'] = all_findings['Z값'].abs()
            top30 = all_findings.nlargest(30, 'Z값_절대').copy()
            
            cols = ['범위', '구분', '첫자리', '관측빈도', '기대빈도', '표본수', 'Z값', '신뢰수준판정']
            return top30[[col for col in cols if col in top30.columns]]
        
        return None
    
    def create_explanation_sheet(self):
        """분석 해설 시트 생성 (세무조사 강화)"""
        explanations = []
        
        # === 종합 평가 ===
        explanations.append({
            '항목': '=== 종합 평가 ===',
            '값': '',
            '평가': '',
            '조치': ''
        })
        
        # 전체 분석 결과
        if '01_벤포드_전체' in self.results:
            df = self.results['01_벤포드_전체']
            
            # Z값 기준 심각도
            red_count = len(df[df['Z값'].abs() >= 2.0])
            orange_count = len(df[(df['Z값'].abs() >= 1.5) & (df['Z값'].abs() < 2.0)])
            yellow_count = len(df[(df['Z값'].abs() >= 1.0) & (df['Z값'].abs() < 1.5)])
            
            severity_score = red_count * 3 + orange_count * 2 + yellow_count * 1
            
            if severity_score > 15:
                risk = "🔴 높음"
                action = "즉각 조사 필요"
            elif severity_score > 8:
                risk = "🟠 중간"
                action = "주의 깊은 검토 필요"
            else:
                risk = "🟡 낮음"
                action = "일반 모니터링"
            
            explanations.append({
                '항목': '종합 위험도',
                '값': f'{severity_score}/27',
                '평가': risk,
                '조치': action
            })
            
            explanations.append({
                '항목': '매우 이상 (|Z|≥2)',
                '값': f'{red_count}개',
                '평가': '🔴',
                '조치': '전수 조사'
            })
            
            explanations.append({
                '항목': '주의 (|Z|≥1.5)',
                '값': f'{orange_count}개',
                '평가': '🟠',
                '조치': '샘플 조사'
            })
            
            explanations.append({
                '항목': '관심 (|Z|≥1)',
                '값': f'{yellow_count}개',
                '평가': '🟡',
                '조치': '모니터링'
            })
            
            # === 가장 이상한 10개 ===
            explanations.append({
                '항목': '',
                '값': '',
                '평가': '',
                '조치': ''
            })
            explanations.append({
                '항목': '=== 가장 이상한 10개 ===',
                '값': '',
                '평가': '',
                '조치': ''
            })
            
            top10 = df.nlargest(10, 'Z값')
            for i, (_, row) in enumerate(top10.iterrows(), 1):
                digit = int(row['첫자리'])
                obs = int(row['관측빈도'])
                exp = row['기대빈도']
                z = row['Z값']
                diff = obs - exp
                pct_diff = (diff / exp * 100) if exp > 0 else 0
                
                explanations.append({
                    '항목': f'{i}. {digit}으로 시작',
                    '값': f'{obs}건 (정상:{exp:.0f}건)',
                    '평가': f'Z={z:.1f}',
                    '조치': f'{diff:+.0f}건 ({pct_diff:+.1f}%)'
                })
        
        # === 데이터 통계 ===
        explanations.append({
            '항목': '',
            '값': '',
            '평가': '',
            '조치': ''
        })
        explanations.append({
            '항목': '=== 데이터 통계 ===',
            '값': '',
            '평가': '',
            '조치': ''
        })
        
        total_entries = len(self.data)
        total_amount = self.data['금액'].sum()
        
        explanations.append({
            '항목': '총 전표 수',
            '값': f'{total_entries:,}개',
            '평가': '',
            '조치': ''
        })
        
        explanations.append({
            '항목': '총 금액',
            '값': f'{total_amount:,.0f}원',
            '평가': '',
            '조치': f'건당 평균 {total_amount/total_entries:,.0f}원'
        })
        
        if '일자' in self.data.columns:
            unique_dates = self.data['일자'].nunique()
            explanations.append({
                '항목': '분석 기간',
                '값': f'{unique_dates}일',
                '평가': '',
                '조치': f'일평균 {total_entries/unique_dates:.0f}건'
            })
        
        if '거래처' in self.data.columns:
            unique_vendors = self.data['거래처'].nunique()
            explanations.append({
                '항목': '총 거래처 수',
                '값': f'{unique_vendors:,}개',
                '평가': '',
                '조치': f'거래처당 평균 {total_entries/unique_vendors:.1f}건'
            })
        
        # === 세무조사 착안점 ===
        explanations.append({
            '항목': '',
            '값': '',
            '평가': '',
            '조치': ''
        })
        explanations.append({
            '항목': '=== 세무조사 착안점 ===',
            '값': '',
            '평가': '',
            '조치': ''
        })
        
        # 1. 라운드 넘버 분석
        if '금액' in self.data.columns:
            round_10k = len(self.data[self.data['금액'] % 10000 == 0])
            round_50k = len(self.data[self.data['금액'] % 50000 == 0])
            round_100k = len(self.data[self.data['금액'] % 100000 == 0])
            
            round_pct = round_10k / total_entries * 100
            
            explanations.append({
                '항목': '라운드 넘버',
                '값': f'{round_10k}건 ({round_pct:.1f}%)',
                '평가': '10,000원 단위',
                '조치': '경비 쪼개기 의심' if round_pct > 20 else '정상 범위'
            })
        
        # 2. 고액 거래 분석
        high_amount = self.data[self.data['금액'] >= 10000000]  # 1천만원 이상
        if len(high_amount) > 0:
            explanations.append({
                '항목': '고액 거래 (≥1천만)',
                '값': f'{len(high_amount)}건',
                '평가': f'{high_amount["금액"].sum():,.0f}원',
                '조치': '증빙서류 중점 확인'
            })
        
        # 3. 손익별 분석
        if '손익' in self.data.columns:
            pnl_summary = self.data.groupby('손익')['금액'].agg(['count', 'sum'])
            for pnl, row in pnl_summary.iterrows():
                explanations.append({
                    '항목': f'{pnl} 거래',
                    '값': f'{int(row["count"]):,}건',
                    '평가': f'{row["sum"]:,.0f}원',
                    '조치': f'평균 {row["sum"]/row["count"]:,.0f}원'
                })
        
        # 4. 상위 거래처 분석
        if '거래처' in self.data.columns and '07_거래처별요약' in self.results:
            top_vendors = self.results['07_거래처별요약'].head(5)
            
            explanations.append({
                '항목': '',
                '값': '',
                '평가': '',
                '조치': ''
            })
            explanations.append({
                '항목': '상위 5대 거래처',
                '값': '',
                '평가': '',
                '조치': ''
            })
            
            for i, (_, row) in enumerate(top_vendors.iterrows(), 1):
                vendor = row['거래처']
                count = row['전표갯수']
                amount = row['금액합계']
                pct = amount / total_amount * 100
                
                explanations.append({
                    '항목': f'{i}. {vendor}',
                    '값': f'{int(count)}건',
                    '평가': f'{amount:,.0f}원',
                    '조치': f'전체의 {pct:.1f}%'
                })
        
        # === 조사 권고사항 ===
        explanations.append({
            '항목': '',
            '값': '',
            '평가': '',
            '조치': ''
        })
        explanations.append({
            '항목': '=== 조사 권고사항 ===',
            '값': '',
            '평가': '',
            '조치': ''
        })
        
        # Z값 기준 권고
        if '01_벤포드_전체' in self.results:
            df = self.results['01_벤포드_전체']
            high_z = df[df['Z값'].abs() >= 3.0]
            
            if len(high_z) > 0:
                explanations.append({
                    '항목': '1. 최우선 조사',
                    '값': f'|Z|≥3인 {len(high_z)}개 숫자',
                    '평가': '통계적으로 거의 불가능',
                    '조치': '해당 금액 전수 조사'
                })
            
            medium_z = df[(df['Z값'].abs() >= 2.0) & (df['Z값'].abs() < 3.0)]
            if len(medium_z) > 0:
                explanations.append({
                    '항목': '2. 높은 우선순위',
                    '값': f'2≤|Z|<3인 {len(medium_z)}개 숫자',
                    '평가': '95% 신뢰구간 밖',
                    '조치': '샘플링 조사'
                })
            
            explanations.append({
                '항목': '3. 증빙 확인',
                '값': '고액 거래 + 상위 거래처',
                '평가': '집중도 높은 항목',
                '조치': '계약서, 세금계산서 확인'
            })
            
            explanations.append({
                '항목': '4. 패턴 분석',
                '값': '라운드 넘버 다수 발견 시',
                '평가': '경비 쪼개기 의심',
                '조치': '적요 + 일자 교차 검증'
            })
        
        return pd.DataFrame(explanations)
    
    def apply_conditional_formatting(self, ws, start_row, end_row, z_col_idx):
        """조건부 서식 적용"""
        colors = {
            'red': PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid'),
            'orange': PatternFill(start_color='FFCC66', end_color='FFCC66', fill_type='solid'),
            'yellow': PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        }
        
        for row_idx in range(start_row, end_row + 1):
            z_cell = ws.cell(row_idx, z_col_idx)
            if z_cell.value is not None:
                try:
                    z_val = abs(float(z_cell.value))
                    
                    if z_val >= 2.0:
                        fill = colors['red']
                    elif z_val >= 1.5:
                        fill = colors['orange']
                    elif z_val >= 1.0:
                        fill = colors['yellow']
                    else:
                        continue
                    
                    # 전체 행에 색상 적용
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row_idx, col_idx).fill = fill
                except:
                    pass
    
    def save_results(self):
        """결과를 엑셀로 저장"""
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 원본 데이터 먼저 (첫자리 컬럼 제거)
            original_data = self.data.drop(columns=['첫자리'], errors='ignore')
            original_data.to_excel(writer, sheet_name='00_원본데이터', index=False)
            
            # 분석 결과들
            for sheet_name, df in self.results.items():
                if df is not None and len(df) > 0:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 조건부 서식 적용
        output.seek(0)
        wb = load_workbook(output)
        
        # 벤포드 시트들에 색상 적용
        for sheet_name in wb.sheetnames:
            if '벤포드' in sheet_name:
                ws = wb[sheet_name]
                if ws.max_row > 1:
                    # Z값 컬럼 찾기
                    z_col_idx = None
                    for col_idx in range(1, ws.max_column + 1):
                        if ws.cell(1, col_idx).value == 'Z값':
                            z_col_idx = col_idx
                            break
                    
                    if z_col_idx:
                        self.apply_conditional_formatting(ws, 2, ws.max_row, z_col_idx)
            
            # 중점검토 시트도
            if sheet_name == '09_중점검토':
                ws = wb[sheet_name]
                if ws.max_row > 1:
                    for col_idx in range(1, ws.max_column + 1):
                        if ws.cell(1, col_idx).value == 'Z값':
                            self.apply_conditional_formatting(ws, 2, ws.max_row, col_idx)
                            break
        
        # 메모리에 저장
        final_output = io.BytesIO()
        wb.save(final_output)
        final_output.seek(0)
        
        return final_output


def main():
    """메인 앱"""
    
    # 제목
    st.title("🔍 벤포드 분석 도구")
    st.markdown("---")
    
    # 사이드바 - 설명
    with st.sidebar:
        st.header("📖 사용 방법")
        st.markdown("""
        ### 1️⃣ 파일 준비
        - '벤포드_분개장_틀.xlsx'에 데이터 입력
        - 필수 컬럼: 일자, 손익, 금액, 거래처, 적요
        
        ### 2️⃣ 파일 업로드
        - 아래에서 파일 선택
        
        ### 3️⃣ 분석 실행
        - 자동으로 분석 시작
        
        ### 4️⃣ 결과 다운로드
        - '벤포드 분석결과.xlsx' 다운로드
        """)
        
        st.markdown("---")
        st.markdown("### 🎯 Z값 의미")
        st.markdown("""
        - **|Z| < 1.0**: 정상 범위
        - **|Z| ≥ 1.0**: 🟡 약간 주의
        - **|Z| ≥ 1.5**: 🟠 주의 필요
        - **|Z| ≥ 2.0**: 🔴 조사 필요
        - **|Z| ≥ 3.0**: 🔴🔴 즉각 조사
        """)
        
        st.markdown("---")
        st.markdown("### 📊 벤포드 법칙")
        st.markdown("""
        자연스러운 숫자 데이터의 **첫 자리**는:
        - **1로 시작**: 30.1% (가장 많음)
        - **2로 시작**: 17.6%
        - **3로 시작**: 12.5%
        - ...
        - **9로 시작**: 4.6% (가장 적음)
        
        조작된 데이터는 이 패턴을 따르지 않습니다.
        """)
    
    # 메인 영역
    st.header("📁 파일 업로드")
    
    # 벤포드 법칙 설명 (Expander)
    with st.expander("📚 벤포드 법칙이란? (클릭하여 펼치기)"):
        st.markdown("""
        ### 🎯 벤포드 법칙 (Benford's Law)
        
        **1938년 물리학자 프랭크 벤포드(Frank Benford)**가 발견한 법칙으로, 
        자연스럽게 발생한 숫자 데이터의 **첫 자리**는 균등하게 분포하지 않고 
        특정한 로그 패턴을 따른다는 통계 법칙입니다.
        
        #### 📊 첫 자리 출현 확률:
        
        | 첫자리 | 확률 | 예시 |
        |--------|------|------|
        | 1 | 30.1% | 1,234원, 10,000원, 157,890원 |
        | 2 | 17.6% | 2,345원, 23,456원, 289,012원 |
        | 3 | 12.5% | 3,456원, 34,567원, 345,678원 |
        | 4 | 9.7% | 4,567원, 45,678원, 456,789원 |
        | 5 | 7.9% | 5,678원, 56,789원, 567,890원 |
        | 6 | 6.7% | 6,789원, 67,890원, 678,901원 |
        | 7 | 5.8% | 7,890원, 78,901원, 789,012원 |
        | 8 | 5.1% | 8,901원, 89,012원, 890,123원 |
        | 9 | 4.6% | 9,012원, 90,123원, 901,234원 |
        

        #### ✅ 적용 조건
        
        벤포드 법칙은 다음 조건에서 잘 작동:
        - 여러 자릿수의 숫자 (multiple orders of magnitude)
        - 자연스럽게 발생한 데이터 (인위적 개입 없음)
        - 큰 금액 (100만원 이상에서 특히 유효)
        - 외주비, 컨설팅비, 매출 등
        
        #### ❌ 부적합한 경우
        
        - 정액제 경비 (식대 9,000원, 교통비 5만원)
        - 계약 금액 (월세 100만원, 리스료)
        - 소액 거래 (1만원 미만)
        - 가격 정책 (9,900원 상품가)
        """)
    
    # 세무조사 활용 설명 (Expander)
    with st.expander("🔍 세무조사에서의 활용 (클릭하여 펼치기)"):
        st.markdown("""
        ### 🎯 세무조사에서 벤포드 법칙 활용
        
        #### 1️⃣ 초기 스크리닝 도구
        
        벤포드 분석은 **대량의 거래 데이터에서 의심 거래를 빠르게 선별**하는 
        초기 스크리닝 도구로 활용됩니다.
        
        **활용 절차:**
        ```
        1단계: 전체 분개장 데이터 벤포드 분석
           ↓
        2단계: Z값 높은 항목 추출 (의심 항목)
           ↓
        3단계: 해당 금액 거래 상세 조사
           ↓
        4단계: 증빙서류 확인 + 담당자 인터뷰
           ↓
        5단계: 부정 여부 판단
        ```
        
        #### 2️⃣ 주요 적발 사례
        
        **🔴 경비 쪼개기 (Split Billing)**
        ```
        정상: 다양한 금액 (12,345원, 67,890원, 234,567원...)
        비정상: 5만원이 100건 (50,000원 × 100)
        
        → 5로 시작하는 금액 비정상적으로 많음
        → 벤포드 분석으로 탐지
        → 가공 경비 적발
        ```
        
        **🔴 매출 누락**
        ```
        정상: 벤포드 패턴 따름
        비정상: 특정 숫자만 누락 (예: 9로 시작하는 매출만 빠짐)
        
        → 벤포드 분포에서 9가 비정상적으로 적음
        → 의도적 매출 누락 의심
        → 탈루 세금 추징
        ```
        
        **🔴 허위 거래처**
        ```
        정상: 거래처마다 다양한 금액
        비정상: 특정 거래처가 라운드 넘버만 거래
        
        → 해당 거래처 벤포드 분석 시 이상
        → 허위 거래처 적발
        → 부정 환급 추징
        ```
        
        #### 3️⃣ 세무조사 착안점
        
        벤포드 분석 결과에서 다음 사항을 중점 확인:
        
        **✅ 1. Z값 기준 우선순위**
        - |Z| ≥ 3.0: 최우선 조사 (통계적으로 거의 불가능)
        - |Z| ≥ 2.0: 높은 우선순위 (95% 신뢰구간 밖)
        - |Z| ≥ 1.5: 중간 우선순위 (샘플링 조사)
        
        **✅ 2. 라운드 넘버 집중도**
        - 10,000원, 50,000원, 100,000원 단위가 20% 이상
        - → 경비 쪼개기 의심
        
        **✅ 3. 특정 계정과목**
        - 접대비, 복리후생비, 여비교통비
        - → 가공 경비 빈발 계정
        
        **✅ 4. 고액 거래**
        - 1천만원 이상 거래
        - → 증빙서류 필수 확인
        
        **✅ 5. 집중 거래처**
        - 상위 5개 거래처가 전체 거래액의 50% 이상
        - → 특수관계 여부 확인
        
        #### 4️⃣ 국세청 활용 사례
        
        **미국 IRS (국세청)**
        - 1990년대부터 세무조사에 벤포드 분석 활용
        - 대규모 탈세 사건 적발에 효과적
        - 디지털 포렌식의 핵심 도구
        
        **한국 국세청**
        - 전산 데이터 분석 시스템에 적용
        - 고액 자산가 및 법인 세무조사 선정 기준
        - 정밀 조사 대상 선별 도구
        
        #### 5️⃣ 법정 증거 능력
        
        ⚠️ **중요:** 벤포드 분석 결과 자체는 **직접 증거가 아님**
        
        - ✅ **가능**: 의심 거래 발견 → 추가 조사의 단서
        - ❌ **불가**: 벤포드 이상 → 즉시 과세 (증빙 필요)
        
        **올바른 활용:**
        1. 벤포드로 의심 항목 선별
        2. 증빙서류 확인 (세금계산서, 계약서 등)
        3. 담당자 인터뷰
        4. 통장 내역 대조
        5. 종합 판단 후 과세
        
        #### 6️⃣ 주의사항
        
        **높은 Z값 ≠ 무조건 부정**
        
        정당한 사유가 있을 수 있음:
        - ✅ 정액제 경비 (단체 식대, 통신비)
        - ✅ 계약 기반 금액 (임대료, 용역비)
        - ✅ 가격 정책 (9,900원 제품가)
        - ✅ 업종 특성 (특정 금액대 집중)
        
        **→ 반드시 원인 확인 후 판단!**
        
        #### 7️⃣ 실무 체크리스트
        
        벤포드 분석 후 확인할 사항:
        
        - [ ] Z값 상위 30개 항목 리스트업
        - [ ] 해당 금액 전표 전수 추출
        - [ ] 거래처별 집계 (손익 구분)
        - [ ] 적요 키워드 분석
        - [ ] 증빙서류 비치 확인
        - [ ] 라운드 넘버 비율 계산
        - [ ] 시계열 패턴 분석 (월별 추이)
        - [ ] 동종 업계 비교
        - [ ] 담당자 면담 준비
        - [ ] 추가 자료 요청 리스트 작성
        
        #### 📚 참고 문헌
        
        - Nigrini, M. (2012). "Benford's Law: Applications for Forensic Accounting, Auditing, and Fraud Detection"
        - Durtschi, C. et al. (2004). "The Effective Use of Benford's Law in Detecting Fraud in Accounting Data"
        - 국세청 세무조사 실무 지침서
        """)
    
    uploaded_file = st.file_uploader(
        "벤포드 분개장 데이터 파일을 선택하세요 (.xlsx)",
        type=['xlsx'],
        help="'벤포드_분개장_틀.xlsx'에 데이터를 입력한 파일"
    )
    
    if uploaded_file is not None:
        try:
            # 파일 정보
            st.success(f"✅ 파일 업로드 완료: {uploaded_file.name}")
            
            # 미리보기
            with st.expander("📋 데이터 미리보기"):
                df_preview = pd.read_excel(uploaded_file, sheet_name='분개장입력', nrows=10)
                st.dataframe(df_preview)
                st.info(f"총 {len(pd.read_excel(uploaded_file, sheet_name='분개장입력')):,}행")
            
            # 분석 버튼
            if st.button("🚀 벤포드 분석 시작", type="primary"):
                st.markdown("---")
                st.header("⚙️ 분석 중...")
                
                # 분석 실행
                analyzer = BenfordAnalyzer(uploaded_file)
                
                with st.spinner("데이터 로딩 중..."):
                    analyzer.load_data()
                
                st.success(f"✅ 데이터 로딩 완료: {len(analyzer.data):,}건")
                
                # 분석 실행
                analyzer.run_analysis()
                
                # 결과 저장
                with st.spinner("결과 파일 생성 중..."):
                    result_file = analyzer.save_results()
                
                st.markdown("---")
                st.header("✅ 분석 완료!")
                
                # 간단한 요약
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.metric("총 전표 수", f"{len(analyzer.data):,}건")
                
                with col2:
                    if '일자' in analyzer.data.columns:
                        days = analyzer.data['일자'].nunique()
                        st.metric("분석 기간", f"{days}일")
                
                with col3:
                    if '거래처' in analyzer.data.columns:
                        vendors = analyzer.data['거래처'].nunique()
                        st.metric("거래처 수", f"{vendors:,}개")
                
                # 주요 발견사항
                if '01_벤포드_전체' in analyzer.results:
                    df = analyzer.results['01_벤포드_전체']
                    red_count = len(df[df['Z값'].abs() >= 2.0])
                    
                    if red_count > 10:
                        st.warning(f"⚠️ 매우 이상한 패턴 {red_count}개 발견! 상세 결과 확인 필요")
                    elif red_count > 0:
                        st.info(f"ℹ️ 주의가 필요한 패턴 {red_count}개 발견")
                    else:
                        st.success("✅ 심각한 이상 패턴 없음")
                
                # 다운로드 버튼
                st.download_button(
                    label="📥 벤포드 분석결과.xlsx 다운로드",
                    data=result_file,
                    file_name=f"벤포드_분석결과_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )
                
                st.markdown("---")
                st.info("""
                💡 **다음 단계:**
                1. 다운로드한 파일을 엑셀에서 열기
                2. '10_분석해설' 시트에서 종합 평가 + 세무조사 착안점 확인
                3. '09_중점검토' 시트에서 이상치 Top 30 확인
                4. '06_일자별요약', '07_거래처별요약' 시트에서 패턴 확인
                5. 🔴빨강, 🟠주황 표시된 항목 우선 조사
                
                ℹ️ **첫 자리 분석:**
                - 1,234원 → 1로 시작
                - 98,765원 → 9로 시작
                - 첫 자리(1-9)만 분석하여 더 정확한 결과 제공
                """)
                
        except Exception as e:
            st.error(f"❌ 오류 발생: {str(e)}")
            st.exception(e)
    
    else:
        # 안내 메시지
        st.info("""
        👆 위에서 파일을 선택하세요.
        
        **파일 준비 방법:**
        1. '벤포드_분개장_틀.xlsx' 다운로드
        2. '분개장입력' 시트에 데이터 입력
        3. 여기에 업로드
        """)


if __name__ == "__main__":
    main()
